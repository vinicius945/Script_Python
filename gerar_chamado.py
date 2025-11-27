import openpyxl
from datetime import datetime
import os
import win32com.client


ARQUIVO_MODELO = 'checklist_individual 2.xlsx'
NOME_TECNICO = "Vinicius Prates Altafini"
LOGIN_TECNICO = "vpaltafini"

LINHA_RESOLUCAO = 51
COLUNA_RESOLUCAO = 2


LINHAS_IGNORAR = [19, 20, 24, 25, 37, 38, 40, 41, 50, 51]

def converter_para_pdf(caminho_excel, caminho_pdf):
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(caminho_excel)
        wb.ExportAsFixedFormat(0, caminho_pdf)
        wb.Close(False)
        excel.Quit()
        print(f"PDF gerado: {os.path.basename(caminho_pdf)}")
    except Exception as e:
        print(f"Erro PDF (tente fechar o Excel): {e}")

def criar_checklist():
    if not os.path.exists(ARQUIVO_MODELO):
        print(f"ERRO: Arquivo '{ARQUIVO_MODELO}' não encontrado.")
        input("Enter para sair...")
        return

    print("--- Gerador V8 (Regras de Admin e Local) ---")
    
    # 1. Coleta de Dados
    ticket = input("Ticket: ")
    cliente = input("Cliente: ") 
    local = input("Local (Ex: Home Office / Matriz): ") # Vamos usar isso na linha 23
    nome_colab = input("Nome Colaborador: ")
    login_colab = input("Login Colaborador: ")
    
   
    hostname = input("Qual o Hostname? ")
    
    tem_admin = input("Usuário tem Admin? (S/N): ").strip().upper()
    if tem_admin == 'S':
        obs_admin = "Possui permissão de Admin local"
    else:
        obs_admin = "Não possui Admin (Perfil Padrão)"

    # Colaborador Respondeu?
    status_padrao = "Concluido"
    obs_padrao = "Verificado"
    respondeu = input("Colaborador respondeu? (S/N): ").strip().upper()
    if respondeu == 'N':
        status_padrao = "N/A"
        obs_padrao = "Colaborador não respondeu"
    
    resolucao_texto = input("Resolução do Chamado: ")
    data_hoje = datetime.now().strftime("%d/%m/%Y")

    # 2. Carrega Planilha
    wb = openpyxl.load_workbook(ARQUIVO_MODELO)
    sheet = wb.active

    
    
    REGRAS_ESPECIFICAS = {
        
        18: f"Hostname: {hostname}",
        
        
        22: obs_admin,
        
       
        23: f"Equipamento utilizado em: {local}"
    }

    # 3. Preenche Cabeçalhos
    sheet['B5'] = ticket
    sheet['D5'] = data_hoje
    sheet['B6'] = cliente
    sheet['D6'] = local
    sheet['B9'] = nome_colab
    sheet['E9'] = login_colab
    sheet['B12'] = NOME_TECNICO
    sheet['E12'] = LOGIN_TECNICO
    
    sheet.cell(row=LINHA_RESOLUCAO, column=COLUNA_RESOLUCAO).value = resolucao_texto

    # 4. Preenche Itens
    print("\n--- Processando Linhas ---")
    linhas_validas = []

    for row in range(16, 66):
        if row in LINHAS_IGNORAR:
            continue

        descricao = sheet.cell(row=row, column=2).value
        
        if descricao and isinstance(descricao, str) and len(descricao) > 5:
            
            
            if row in REGRAS_ESPECIFICAS:
                texto_status = "Concluido" 
                texto_obs = REGRAS_ESPECIFICAS[row] 
                print(f"[{row}] CONFIGURADO: {texto_obs}")
            else:
                
                texto_status = status_padrao
                texto_obs = obs_padrao
                print(f"[{row}] Padrão...")

            
            sheet.cell(row=row, column=5).value = texto_status
            sheet.cell(row=row, column=6).value = texto_obs
            
            linhas_validas.append(row)

    # 5. Ajuste Manual
    print("\n--- Deseja ajustar algo manualmente? ---")
    while True:
        opcao = input("Linha (ou ENTER p/ gerar): ")
        if not opcao: break
        
        try:
            linha = int(opcao)
            if linha in linhas_validas:
                print(f"Editando linha {linha}...")
                n_stat = input("Novo Status: ")
                n_obs = input("Nova Obs: ")
                if n_stat: sheet.cell(row=linha, column=5).value = n_stat
                if n_obs: sheet.cell(row=linha, column=6).value = n_obs
            else:
                print("Linha inválida.")
        except ValueError:
            print("Use números.")

    # 6. Salvar
    nome_xlsx = f"Checklist {ticket}.xlsx"
    wb.save(nome_xlsx)
    
    caminho_xlsx = os.path.abspath(nome_xlsx)
    caminho_pdf = os.path.abspath(f"Checklist {ticket}.pdf")
    
    print("Gerando PDF...")
    converter_para_pdf(caminho_xlsx, caminho_pdf)
    
    input("\nFim. Pressione ENTER.")

if __name__ == "__main__":
    criar_checklist()